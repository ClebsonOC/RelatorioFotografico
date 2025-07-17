# report_logic.py
# -*- coding: utf-8 -*-
# Descrição: Módulo com a lógica de negócio para gerar o relatório fotográfico. (v28 - Correção de I/O de imagem)

import openpyxl
from openpyxl.drawing.image import Image as ExcelImage
from openpyxl.utils import get_column_letter
from copy import copy
from pathlib import Path
from datetime import datetime as DatetimeObject
import io
import logging
from collections import defaultdict

try:
    from PIL import Image as PilImage
    PIL_AVAILABLE = True
except ImportError:
    PIL_AVAILABLE = False

# --- CONSTANTES ---
NOME_ABA_MODELO = "MODELO"
CELULA_NOME_RUA = "K13"
TARGET_ABAS_FOTOS = 40
MIN_FOTOS_POR_RUA = 3
PALAVRAS_IGNORADAS = {"TAPA BURACO"}
EXTENSOES_IMAGEM = {'.jpg', '.jpeg', '.png', '.bmp', '.gif', '.tif', '.tiff'}
IMG_ALTURA_PX = 370
IMG_LARGURA_PX = 730
ANCHOR_FOTOS = ["B16", "B39", "B62"]
NOMES_FOTOS_ESPERADOS = ['1', '2', '3'] 

logging.basicConfig(level=logging.INFO, format='%(asctime)s - %(levelname)s - %(message)s')

# --- FUNÇÕES COM REPORT DE STATUS ---

def extrair_dados_medicao(caminho_arquivo_medicao, nome_aba_medicao, status_callback=None):
    """Extrai dados da planilha, lendo Data(ColB), Rua(ColC) e Prioridade(ColF)."""
    def _report(log_type, message, **kwargs):
        if status_callback: status_callback(log_type, message, **kwargs)

    _report("log", f"A abrir planilha de medição: '{caminho_arquivo_medicao.name}'...")
    dados_extraidos = []
    workbook = None
    try:
        workbook = openpyxl.load_workbook(caminho_arquivo_medicao, read_only=True, data_only=True)
        _report("log", f"Planilha aberta. A aceder à aba '{nome_aba_medicao}'...")
        
        if nome_aba_medicao not in workbook.sheetnames:
            _report("error", f"Aba '{nome_aba_medicao}' não encontrada.")
            return None

        sheet = workbook[nome_aba_medicao]
        max_row = sheet.max_row
        total_linhas = max_row - 6 if max_row > 6 else 0
        _report("log", f"A ler {total_linhas} linhas da aba '{nome_aba_medicao}'.")

        for i, row_idx in enumerate(range(7, max_row + 1)):
            if i % 50 == 0 or i == total_linhas -1:
                 _report("progress", f"Lendo planilha ({i+1}/{total_linhas})", value=i + 1, total=total_linhas, step="extract")

            cell_data_valor = sheet.cell(row=row_idx, column=2).value
            cell_rua_valor = sheet.cell(row=row_idx, column=3).value
            
            if not cell_rua_valor or not isinstance(cell_rua_valor, str):
                continue
            
            nome_rua_normalizado = cell_rua_valor.strip()
            if not nome_rua_normalizado or any(p_ign in nome_rua_normalizado.upper() for p_ign in PALAVRAS_IGNORADAS):
                continue
            
            nome_pasta_data_str = None
            if isinstance(cell_data_valor, DatetimeObject):
                nome_pasta_data_str = cell_data_valor.strftime('%d-%m-%Y')
            elif isinstance(cell_data_valor, str) and cell_data_valor.strip():
                try:
                    parsed_date = None
                    for fmt in ('%d-%m-%Y', '%d/%m/%Y', '%Y-%m-%d'):
                        try:
                            parsed_date = DatetimeObject.strptime(cell_data_valor.strip(), fmt)
                            break
                        except ValueError:
                            continue
                    if parsed_date:
                        nome_pasta_data_str = parsed_date.strftime('%d-%m-%Y')
                    else:
                        continue
                except ValueError:
                    continue
            else:
                continue

            cell_prio_valor = sheet.cell(row=row_idx, column=6).value
            prioridade_num = float('-inf')
            if cell_prio_valor is not None:
                try: prioridade_num = float(cell_prio_valor)
                except (ValueError, TypeError): pass

            dados_extraidos.append({'data_pasta': nome_pasta_data_str, 'nome_rua': nome_rua_normalizado, 'prioridade': prioridade_num, 'linha_origem': row_idx})
        
        _report("log", f"Extração de dados da planilha concluída. {len(dados_extraidos)} registos válidos encontrados.")
        return dados_extraidos
    finally:
        if workbook: workbook.close()


def processar_relatorio(caminho_medicao, aba_medicao, dir_fotos, caminho_modelo_excel, status_callback=None):
    """Orquestra todo o processo, garantindo que cada data seja representada."""
    def _report(log_type, message, **kwargs):
        if status_callback: status_callback(log_type, message, **kwargs)

    try:
        _report("log", "Iniciando processo. Etapa 1: Extrair dados da planilha de medição.")
        lista_dados_medicao = extrair_dados_medicao(Path(caminho_medicao), aba_medicao, status_callback)
        
        if not lista_dados_medicao:
            _report("result", "Nenhum dado válido encontrado na planilha de medição.", success=False)
            return

        _report("log", f"Etapa 2: Verificar fotos para {len(lista_dados_medicao)} registros.")
        itens_com_fotos_suficientes = []
        min_fotos_necessarias = min(MIN_FOTOS_POR_RUA, len(NOMES_FOTOS_ESPERADOS))
        total_medicao = len(lista_dados_medicao)

        for i, dado in enumerate(lista_dados_medicao):
            _report("progress", f"Verificando fotos ({i+1}/{total_medicao})", value=i+1, total=total_medicao, step="photos")
            fotos = encontrar_fotos_rua(Path(dir_fotos), dado['data_pasta'], dado['nome_rua'], status_callback)
            if len(fotos) >= min_fotos_necessarias:
                dado['fotos_para_rua'] = fotos
                itens_com_fotos_suficientes.append(dado)
        
        _report("log", f"Verificação de fotos concluída. {len(itens_com_fotos_suficientes)} itens com fotos suficientes encontrados.")
        if not itens_com_fotos_suficientes:
            _report("result", f"Nenhum item atendeu ao critério de ter no mínimo {min_fotos_necessarias} fotos (nomeadas 1, 2, 3).", success=False)
            return

        _report("log", f"Etapa 3: Aplicando lógica de seleção para garantir uma aba por data.")
        
        itens_por_data = defaultdict(list)
        for item in itens_com_fotos_suficientes:
            itens_por_data[item['data_pasta']].append(item)
        _report("log", f"Agrupamento concluído: {len(itens_por_data)} datas diferentes encontradas com fotos.")

        lista_processar_final = []
        linhas_ja_selecionadas = set()
        
        datas_ordenadas = sorted(itens_por_data.keys(), key=lambda d: DatetimeObject.strptime(d, '%d-%m-%Y'))

        _report("log", "Primeira passagem: Garantindo uma rua de maior prioridade por data.")
        for data in datas_ordenadas:
            if len(lista_processar_final) >= TARGET_ABAS_FOTOS:
                break
            
            candidatos_da_data = sorted(itens_por_data[data], key=lambda x: (-x['prioridade'], x['linha_origem']))
            
            melhor_candidato = candidatos_da_data[0]
            lista_processar_final.append(melhor_candidato)
            linhas_ja_selecionadas.add(melhor_candidato['linha_origem'])

        _report("log", f"Após a primeira passagem, {len(lista_processar_final)} abas selecionadas (uma por data).")

        if len(lista_processar_final) < TARGET_ABAS_FOTOS:
            _report("log", f"Segunda passagem: Preenchendo as {TARGET_ABAS_FOTOS - len(lista_processar_final)} vagas restantes.")
            
            candidatos_restantes = [
                item for item in itens_com_fotos_suficientes 
                if item['linha_origem'] not in linhas_ja_selecionadas
            ]
            
            candidatos_restantes.sort(key=lambda x: (-x['prioridade'], x['linha_origem']))

            ruas_adicionadas_na_passagem1 = {item['nome_rua'] for item in lista_processar_final}
            
            for item in candidatos_restantes:
                if len(lista_processar_final) >= TARGET_ABAS_FOTOS:
                    break
                if item['nome_rua'] not in ruas_adicionadas_na_passagem1:
                    lista_processar_final.append(item)
            
            if len(lista_processar_final) < TARGET_ABAS_FOTOS:
                 for item in candidatos_restantes:
                    if len(lista_processar_final) >= TARGET_ABAS_FOTOS:
                        break
                    if not any(item['linha_origem'] == x['linha_origem'] for x in lista_processar_final):
                         lista_processar_final.append(item)

        _report("log", f"Seleção final concluída: {len(lista_processar_final)} itens serão processados.")
        
        lista_processar_final.sort(key=lambda x: x['linha_origem'])

        _report("log", "Etapa 4: Iniciar a criação do ficheiro Excel. Isto pode demorar...")
        sucesso, abas_criadas, arq_final = gerar_arquivo_excel(lista_processar_final, Path(caminho_modelo_excel), status_callback)

        if sucesso:
            _report("result", f"Relatório gerado com {abas_criadas} abas.", success=True, file_path=str(arq_final))
        else:
            _report("result", "Ocorreu um erro ao gerar o ficheiro Excel.", success=False)

    except Exception as e:
        logging.exception("Erro inesperado durante o processamento.")
        _report("result", f"Erro inesperado durante o processamento: {e}", success=False)


def gerar_arquivo_excel(itens_a_processar, caminho_modelo, status_callback=None):
    """Gera o ficheiro .xlsx final, copiando a aba modelo e garantindo que o logo seja replicado."""
    def _report(log_type, message, **kwargs):
        if status_callback: status_callback(log_type, message, **kwargs)

    final_workbook = None
    try:
        _report("log", "A carregar o ficheiro modelo para a memória...")
        final_workbook = openpyxl.load_workbook(caminho_modelo)

        if NOME_ABA_MODELO not in final_workbook.sheetnames:
            _report("error", f"A aba modelo '{NOME_ABA_MODELO}' não foi encontrada no ficheiro.")
            return False, 0, None
        
        template_sheet = final_workbook[NOME_ABA_MODELO]
        
        # --- LÓGICA PARA COPIAR O LOGO ---
        logo_info = None
        # CORREÇÃO: O atributo correto é _images, com underscore.
        if template_sheet._images:
            try:
                # Pega a primeira imagem da aba modelo (assumindo que é o logo)
                logo_original = template_sheet._images[0]
                logo_bytes = logo_original._data()
                logo_info = {
                    # CORREÇÃO: Armazena os bytes brutos, não um stream.
                    "data": logo_bytes,
                    "anchor": logo_original.anchor,
                    "width": logo_original.width,
                    "height": logo_original.height
                }
                _report("log", f"Logo encontrado na aba modelo na posição {logo_info['anchor']}. Será replicado.")
            except Exception as e:
                _report("error", f"Não foi possível extrair a imagem da aba modelo. Erro: {e}")
        else:
            _report("log", "Nenhuma imagem (logo) encontrada na aba modelo para replicar.")
        # --- FIM DA LÓGICA DO LOGO ---
        
        total_itens = len(itens_a_processar)
        abas_criadas = 0

        for i, item in enumerate(itens_a_processar):
            _report("progress", f"Criando aba {i+1} de {total_itens}", value=i+1, total=total_itens, step="generate")
            
            new_sheet = final_workbook.copy_worksheet(template_sheet)
            
            nome_aba_unico = f"{item['data_pasta']}_{item['nome_rua']}"
            temp_name = nome_aba_unico[:31]
            count = 1
            while temp_name in final_workbook.sheetnames:
                suffix = f"_{count}"
                temp_name = f"{nome_aba_unico[:31-len(suffix)]}{suffix}"
                count += 1
            new_sheet.title = temp_name

            new_sheet[CELULA_NOME_RUA] = item['nome_rua']
            
            # --- REINSERE O LOGO NA NOVA ABA ---
            if logo_info:
                try:
                    # CORREÇÃO: Cria um novo stream a partir dos bytes para cada imagem.
                    logo_stream = io.BytesIO(logo_info['data'])
                    img_logo = ExcelImage(logo_stream)
                    img_logo.width = logo_info['width']
                    img_logo.height = logo_info['height']
                    new_sheet.add_image(img_logo, logo_info['anchor'])
                except Exception as e:
                    _report("error", f"Falha ao reinserir o logo na aba '{new_sheet.title}': {e}")
            # --- FIM DA REINSERÇÃO DO LOGO ---

            for idx, foto_path in enumerate(item['fotos_para_rua']):
                if idx >= len(ANCHOR_FOTOS): break
                try:
                    with open(foto_path, 'rb') as f_img:
                        stream = io.BytesIO(f_img.read())
                    img_obj = ExcelImage(stream)
                    img_obj.height = IMG_ALTURA_PX
                    img_obj.width = IMG_LARGURA_PX
                    new_sheet.add_image(img_obj, ANCHOR_FOTOS[idx])
                except Exception as e:
                    _report("error", f"Falha ao adicionar foto '{foto_path.name}' à aba '{new_sheet.title}': {e}")
            
            abas_criadas += 1

        if abas_criadas > 0 and NOME_ABA_MODELO in final_workbook.sheetnames:
            _report("log", "A remover a aba 'MODELO' original.")
            final_workbook.remove(template_sheet)
        
        ts = DatetimeObject.now().strftime("%Y%m%d_%H%M%S")
        nome_out = f"Relatorio_Fotografico_{ts}.xlsx"
        novo_path = caminho_modelo.parent / nome_out

        _report("log", "A guardar o ficheiro final. Esta etapa pode ser demorada...")
        final_workbook.save(novo_path)
        _report("log", f"Relatório guardado em: '{novo_path}'")
        return True, abas_criadas, novo_path

    except Exception as e:
        logging.exception("Erro fatal na geração do Excel.")
        _report("error", f"Erro fatal na geração do Excel: {e}")
        return False, 0, None
    finally:
        if final_workbook:
            final_workbook.close()


def encontrar_fotos_rua(base_dir_fotos, nome_pasta_data, nome_rua, status_callback=None):
    """Procura especificamente por fotos nomeadas '1', '2', '3'."""
    def _report(log_type, message, **kwargs):
        if status_callback and log_type != 'debug': status_callback(log_type, message, **kwargs)

    pasta_data_especifica = Path(base_dir_fotos) / nome_pasta_data
    fotos_encontradas = []

    if not pasta_data_especifica.is_dir():
        return fotos_encontradas

    pasta_rua_encontrada = None
    nome_rua_lower = nome_rua.lower()
    try:
        for item_data in pasta_data_especifica.iterdir():
            if item_data.is_dir() and item_data.name.lower() == nome_rua_lower:
                pasta_rua_encontrada = item_data
                break
    except OSError:
        return fotos_encontradas
            
    if not pasta_rua_encontrada:
        # Adiciona log se a pasta da rua não for encontrada
        _report("log", f"AVISO: Pasta para a rua '{nome_rua}' não encontrada dentro de '{nome_pasta_data}'.")
        return fotos_encontradas

    for nome_base in NOMES_FOTOS_ESPERADOS:
        foto_encontrada_para_nome_base = None
        for extensao in EXTENSOES_IMAGEM:
            caminho_foto_tentativa = pasta_rua_encontrada / f"{nome_base}{extensao}"
            if caminho_foto_tentativa.is_file():
                valid = True
                if PIL_AVAILABLE:
                    try:
                        with PilImage.open(caminho_foto_tentativa) as img:
                            img.verify()
                        with PilImage.open(caminho_foto_tentativa) as img:
                            img.load()
                    except Exception as e_pil:
                        _report("log", f"AVISO: Imagem '{caminho_foto_tentativa.name}' parece inválida e será ignorada. Erro: {e_pil}")
                        valid = False
                
                if valid:
                    foto_encontrada_para_nome_base = caminho_foto_tentativa
                    break
        
        if foto_encontrada_para_nome_base:
            fotos_encontradas.append(foto_encontrada_para_nome_base)

    return fotos_encontradas
